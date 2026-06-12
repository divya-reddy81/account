from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE = Path(r"c:\Users\DELL\divPSA\docs\CBO_Setup_RPG_ID_Chain_Reset")
BASE.mkdir(parents=True, exist_ok=True)


def write_text(name: str, content: str) -> None:
    (BASE / name).write_text(content, encoding="utf-8")


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)


def add_filter(ws):
    ws.auto_filter.ref = ws.dimensions


executive_md = """# Executive Summary - CBO Setup RPG ID Chain Reset

## What This Automation Does
This screen flow performs a one-time setup action for trial orgs created through a sign-up process. It updates three PSA hierarchy records (Region, Practice, Group) so their reversed ID chain value becomes `1`, then renames a placeholder PSA resource contact (`Zack Mundy`) to the currently logged-in user.

## Why It Exists
During trial org provisioning, records may still reference the original source chain from the Trial Source Org (TSO). This flow localizes that chain and maps seeded project assignments to the real trial user so "My Projects" user experiences work immediately.

## Business Process Supported
- Trial org post-provisioning setup
- Initial data alignment for PSA hierarchy
- Initial user assignment visibility in project-centric LWCs

## Who Uses It
- Primarily implementation/setup users in newly created sign-up orgs
- Can be initiated by a Salesforce user with permission to run flows and update affected records

## When It Runs
- Manually, from a screen flow launch context
- Intended only for sign-up-generated orgs (explicitly not for Sandbox/Production, as documented in the flow description)

## Expected Business Outcome
- RPG chain reset values are standardized to local org expectations
- Trial user sees project data via components filtered by "My Projects"
- Setup team can confirm completion through success message on screen
"""


technical_md = """# Technical Documentation - CBO Setup RPG ID Chain Reset

## Component Profile
| Item | Value |
|---|---|
| Component Type | Screen Flow |
| API Name | `CBO_Setup_RPG_ID_Chain_Reset` |
| File Name | `force-app/main/default/flows/CBO_Setup_RPG_ID_Chain_Reset.flow-meta.xml` |
| Status | Active |
| API Version | 65.0 |
| Process Type | Flow |
| Start Type | Screen flow start node with direct connector |

## Entry Points and Invocation
| Area | Detail |
|---|---|
| Entry Point | Flow Start element |
| Trigger Event | Manual user launch |
| Invocation Source | UI flow launch (e.g., app/page/button embedding the flow) |
| Run Context | Current running user |
| User Inputs | None (no form inputs); only success display screen |

## Execution Path
1. `Start` -> `Update_Region_ID_Chain`
2. `Update_Region_ID_Chain` updates `pse__Region__c` where `Name = Global Region`
3. `Update_Practice_ID_Chain` updates `pse__Practice__c` where `Name = Global Practice`
4. `Update_Group_ID_Chain` updates `pse__Grp__c` where `Name = Global Group`
5. `Update_Zack_Mundy_to_Running_User` updates `Contact` where `FirstName = Zack` and `LastName = Mundy`
6. `Reversed_RPG_Chain_IDS` screen shows success message

## Related Metadata and Dependencies
| Type | Name/API | Dependency Nature |
|---|---|---|
| Flow | `CBO_Setup_RPG_ID_Chain_Reset` | Primary automation |
| Object | `pse__Region__c` | Updated record target |
| Object | `pse__Practice__c` | Updated record target |
| Object | `pse__Grp__c` | Updated record target |
| Object | `Contact` | Updated record target |
| Field | `pse__Reversed_ID_Chain__c` | Set to literal `1` on three objects |
| Global Variable | `$User.FirstName`, `$User.LastName` | Used for name replacement |
| External Package | PSA namespace (`pse__`) | Managed package object model dependency |

## Reusable Utilities Used
- None explicitly called (no Apex actions/subflows/invocable actions in this flow)

## Integration Endpoints
- No external APIs/endpoints are called by this flow

## Governor Limit Considerations
- Low SOQL/DML complexity due to direct `Record Update` elements with narrow filters
- Still sensitive to record cardinality risk: if filter criteria are not unique, multiple records can be updated unexpectedly
- No loop constructs; therefore low risk of loop-amplified limits

## Security Considerations
- Flow executes in user context (subject to runtime behavior and org settings)
- Requires object/field update access on PSA hierarchy objects and Contact
- Potential data integrity risk from hardcoded person-name matching (`Zack Mundy`)
- No explicit CRUD/FLS enforcement logic because this is declarative flow configuration

## Plain-English Business Meaning of Key Technical Terms
| Technical Term | Simple Explanation | Business Meaning | Example User Action |
|---|---|---|---|
| Reversed ID Chain | A hierarchy tracking value stored on PSA records | Makes hierarchy behave like a local trial org | Admin runs setup flow after sign-up |
| Record Update | Salesforce action that edits existing records | Applies required setup values automatically | Click "Run Flow" once in setup app |
| Running User | Person currently logged in | Determines who gets mapped to seeded resource assignments | New trial user launches the flow |
| Managed Package Object | Standardized object delivered by installed app (PSA) | Enables PSA hierarchy and project operations | PSA app installed in org |
"""


mermaid_md = """# Mermaid Diagrams - CBO Setup RPG ID Chain Reset

## End-to-End Process Flow
```mermaid
flowchart TD
    A[Start Flow] --> B[Update Region ID Chain\\nObject: pse__Region__c\\nFilter: Name = Global Region\\nSet Reversed_ID_Chain = 1]
    B --> C[Update Practice ID Chain\\nObject: pse__Practice__c\\nFilter: Name = Global Practice\\nSet Reversed_ID_Chain = 1]
    C --> D[Update Group ID Chain\\nObject: pse__Grp__c\\nFilter: Name = Global Group\\nSet Reversed_ID_Chain = 1]
    D --> E[Update Zack Mundy Contact\\nObject: Contact\\nFilter: FirstName=Zack AND LastName=Mundy\\nSet names from $User]
    E --> F[Success Screen]
```

## Sequence Diagram
```mermaid
sequenceDiagram
    participant U as Setup User
    participant F as Flow: CBO_Setup_RPG_ID_Chain_Reset
    participant R as pse__Region__c
    participant P as pse__Practice__c
    participant G as pse__Grp__c
    participant C as Contact

    U->>F: Launch flow
    F->>R: Update Reversed_ID_Chain__c = 1 where Name='Global Region'
    F->>P: Update Reversed_ID_Chain__c = 1 where Name='Global Practice'
    F->>G: Update Reversed_ID_Chain__c = 1 where Name='Global Group'
    F->>C: Update FirstName/LastName from running user where name is Zack Mundy
    F-->>U: Display success message
```

## Dependency Graph
```mermaid
graph LR
    Flow[CBO_Setup_RPG_ID_Chain_Reset] --> O1[pse__Region__c]
    Flow --> O2[pse__Practice__c]
    Flow --> O3[pse__Grp__c]
    Flow --> O4[Contact]
    Flow --> GV[$User.FirstName / $User.LastName]
    O1 --> F1[pse__Reversed_ID_Chain__c]
    O2 --> F1
    O3 --> F1
```
"""


flowcharts_md = """# Automation Flowcharts - CBO Setup RPG ID Chain Reset

## Trigger Execution Flow
```mermaid
flowchart LR
    T[Manual Launch] --> S[Start]
    S --> U1[Region Update]
    U1 --> U2[Practice Update]
    U2 --> U3[Group Update]
    U3 --> U4[Contact Rename]
    U4 --> End[Completion Screen]
```

## Decision/Validation Representation
The flow has no explicit Decision elements, but each Record Update contains implicit filter checks:
- Region update only executes against records where `Name = Global Region`
- Practice update only executes against records where `Name = Global Practice`
- Group update only executes against records where `Name = Global Group`
- Contact rename only executes where `FirstName = Zack` and `LastName = Mundy`

If no records match, the flow continues without explicit fault branching.
"""


write_text("Executive_Summary.md", executive_md)
write_text("Technical_Documentation.md", technical_md)
write_text("Mermaid_Diagrams.md", mermaid_md)
write_text("Automation_Flowcharts.md", flowcharts_md)


def make_inventory():
    wb = Workbook()
    ws = wb.active
    ws.title = "Automation Inventory"
    ws.append(
        [
            "Automation Name",
            "Type",
            "API Name",
            "Status",
            "Business Purpose",
            "Invocation",
            "Primary Users",
            "Complexity",
            "Risk Level",
            "Developer Notes",
        ]
    )
    ws.append(
        [
            "CBO Setup RPG ID Chain Reset / Rename Zack Mundy as Running User",
            "Screen Flow",
            "CBO_Setup_RPG_ID_Chain_Reset",
            "Active",
            "Reset RPG chain values and map seeded resource to running trial user",
            "Manual launch",
            "Setup/implementation users in sign-up orgs",
            "Low-Medium",
            "Medium",
            "Not intended for Sandbox/Production per flow description",
        ]
    )
    style_header(ws)
    add_filter(ws)
    auto_width(ws)
    risk_fill = PatternFill("solid", fgColor="FFD966")
    ws["I2"].fill = risk_fill
    wb.save(BASE / "Automation_Inventory.xlsx")


def make_object_field_impact():
    wb = Workbook()
    ws = wb.active
    ws.title = "Object Field Impact"
    ws.append(
        [
            "Object Name",
            "Field API Name",
            "Field Label",
            "Operation Type",
            "Usage Classification",
            "Validation Dependency",
            "Formula Dependency",
            "Lookup/MD Dependency",
            "Related Objects",
            "Decision Driver",
            "BA-Friendly Description",
            "Developer Notes",
        ]
    )
    rows = [
        [
            "pse__Region__c",
            "Name",
            "Region Name",
            "Read",
            "Referenced",
            "Unknown",
            "Unknown",
            "Likely parent hierarchy dependency",
            "pse__Practice__c, pse__Grp__c",
            "Yes",
            "Identifies the specific global region record to be reset.",
            "Filter criterion equals 'Global Region'.",
        ],
        [
            "pse__Region__c",
            "pse__Reversed_ID_Chain__c",
            "Reversed ID Chain",
            "Update",
            "Updated",
            "Unknown",
            "Potential downstream formula/report usage",
            "Hierarchy behavior dependency",
            "pse__Practice__c",
            "No",
            "Sets the hierarchy-chain indicator to local default value.",
            "Set to literal '1'.",
        ],
        [
            "pse__Practice__c",
            "Name",
            "Practice Name",
            "Read",
            "Referenced",
            "Unknown",
            "Unknown",
            "Likely child of region",
            "pse__Region__c, pse__Grp__c",
            "Yes",
            "Finds the global practice record for reset.",
            "Filter criterion equals 'Global Practice'.",
        ],
        [
            "pse__Practice__c",
            "pse__Reversed_ID_Chain__c",
            "Reversed ID Chain",
            "Update",
            "Updated",
            "Unknown",
            "Potential downstream formula/report usage",
            "Hierarchy behavior dependency",
            "pse__Region__c, pse__Grp__c",
            "No",
            "Applies local chain-reset value to global practice.",
            "Set to literal '1'.",
        ],
        [
            "pse__Grp__c",
            "Name",
            "Group Name",
            "Read",
            "Referenced",
            "Unknown",
            "Unknown",
            "Likely child of practice",
            "pse__Practice__c",
            "Yes",
            "Finds the global group record for reset.",
            "Filter criterion equals 'Global Group'.",
        ],
        [
            "pse__Grp__c",
            "pse__Reversed_ID_Chain__c",
            "Reversed ID Chain",
            "Update",
            "Updated",
            "Unknown",
            "Potential downstream formula/report usage",
            "Hierarchy behavior dependency",
            "pse__Practice__c",
            "No",
            "Applies local chain-reset value to global group.",
            "Set to literal '1'.",
        ],
        [
            "Contact",
            "FirstName",
            "First Name",
            "Read + Update",
            "Decision Driver + Updated",
            "Unknown",
            "Unknown",
            "Potential PSA resource mapping dependency",
            "User",
            "Yes",
            "Finds placeholder contact then changes first name to current user first name.",
            "Filter: FirstName='Zack'; value from $User.FirstName.",
        ],
        [
            "Contact",
            "LastName",
            "Last Name",
            "Read + Update",
            "Decision Driver + Updated",
            "Unknown",
            "Unknown",
            "Potential PSA resource mapping dependency",
            "User",
            "Yes",
            "Finds placeholder contact then changes last name to current user last name.",
            "Filter: LastName='Mundy'; value from $User.LastName.",
        ],
        [
            "User (Global Variable)",
            "$User.FirstName",
            "Running User First Name",
            "Read",
            "Referenced",
            "N/A",
            "N/A",
            "N/A",
            "Contact",
            "No",
            "Source value used to rename placeholder resource.",
            "Runtime context variable.",
        ],
        [
            "User (Global Variable)",
            "$User.LastName",
            "Running User Last Name",
            "Read",
            "Referenced",
            "N/A",
            "N/A",
            "N/A",
            "Contact",
            "No",
            "Source value used to rename placeholder resource.",
            "Runtime context variable.",
        ],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    add_filter(ws)
    auto_width(ws)
    wb.save(BASE / "Object_Field_Impact.xlsx")


def make_dependency_map():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Flow Dependencies"
    ws1.append(
        [
            "Source Component",
            "Dependency Type",
            "Target",
            "Direction",
            "Criticality",
            "Notes",
            "Link",
        ]
    )
    data = [
        [
            "CBO_Setup_RPG_ID_Chain_Reset",
            "Object Update",
            "pse__Region__c",
            "Flow -> Object",
            "High",
            "Updates reversed chain on global region",
            "Open technical doc",
        ],
        [
            "CBO_Setup_RPG_ID_Chain_Reset",
            "Object Update",
            "pse__Practice__c",
            "Flow -> Object",
            "High",
            "Updates reversed chain on global practice",
            "Open technical doc",
        ],
        [
            "CBO_Setup_RPG_ID_Chain_Reset",
            "Object Update",
            "pse__Grp__c",
            "Flow -> Object",
            "High",
            "Updates reversed chain on global group",
            "Open technical doc",
        ],
        [
            "CBO_Setup_RPG_ID_Chain_Reset",
            "Object Update",
            "Contact",
            "Flow -> Object",
            "High",
            "Renames placeholder resource contact",
            "Open technical doc",
        ],
        [
            "CBO_Setup_RPG_ID_Chain_Reset",
            "Global Variable",
            "$User",
            "Flow <- Context",
            "Medium",
            "Reads running user's first and last name",
            "Open technical doc",
        ],
    ]
    for r in data:
        ws1.append(r)
    style_header(ws1)
    add_filter(ws1)
    auto_width(ws1)
    for i in range(2, 7):
        ws1[f"G{i}"].hyperlink = "Technical_Documentation.md"
        ws1[f"G{i}"].style = "Hyperlink"

    ws2 = wb.create_sheet("Dependency Risk View")
    ws2.append(["Dependency", "Risk Level", "Rationale", "Mitigation"])
    rows = [
        [
            "Hardcoded Name filters (Global Region/Practice/Group)",
            "Medium",
            "Name changes can break intended target matching",
            "Use externalized config (Custom Metadata) or stable IDs",
        ],
        [
            "Hardcoded contact person name Zack Mundy",
            "High",
            "Could match wrong contact or none in changed datasets",
            "Use dedicated marker field or custom metadata mapping",
        ],
        [
            "PSA managed package object dependency",
            "Medium",
            "Package upgrades may alter behavior/field semantics",
            "Regression test after package updates",
        ],
    ]
    for r in rows:
        ws2.append(r)
    style_header(ws2)
    add_filter(ws2)
    auto_width(ws2)
    high = PatternFill("solid", fgColor="F4CCCC")
    med = PatternFill("solid", fgColor="FFE599")
    for row in ws2.iter_rows(min_row=2, max_col=4):
        level = row[1].value
        row[1].fill = high if level == "High" else med
    wb.save(BASE / "Dependency_Map.xlsx")


def make_integration_map():
    wb = Workbook()
    ws = wb.active
    ws.title = "Integration Map"
    ws.append(
        [
            "Integration Point",
            "Type",
            "Endpoint/System",
            "Direction",
            "Used by Flow?",
            "Business Impact",
            "Notes",
        ]
    )
    ws.append(
        [
            "None detected",
            "N/A",
            "N/A",
            "N/A",
            "No",
            "No direct external integration impact",
            "Flow only performs in-org record updates and screen notification",
        ]
    )
    style_header(ws)
    add_filter(ws)
    auto_width(ws)
    wb.save(BASE / "Integration_Map.xlsx")


def make_risks():
    wb = Workbook()
    ws = wb.active
    ws.title = "Risks"
    ws.append(
        [
            "Risk ID",
            "Category",
            "Risk Description",
            "Severity",
            "Likelihood",
            "Impact",
            "Recommendation",
            "Owner",
            "Status",
        ]
    )
    rows = [
        [
            "R-001",
            "Data Integrity",
            "Hardcoded contact matching by name can update unintended person",
            "High",
            "Medium",
            "Incorrect resource assignment visibility",
            "Use a dedicated external ID/flag field for the seeded contact",
            "Salesforce Admin",
            "Open",
        ],
        [
            "R-002",
            "Configuration Drift",
            "Global hierarchy names may differ across environments",
            "Medium",
            "Medium",
            "Chain reset may skip target records",
            "Store target keys in custom metadata and validate presence before update",
            "Solution Architect",
            "Open",
        ],
        [
            "R-003",
            "Security/FLS",
            "Flow may fail for users lacking update permissions",
            "Medium",
            "Medium",
            "Setup failure and partial execution",
            "Restrict run access to setup permission set and add fault screen logging",
            "Admin + Security",
            "Open",
        ],
        [
            "R-004",
            "Operational",
            "Flow intended for sign-up orgs might be run in non-target orgs",
            "High",
            "Low",
            "Unexpected data changes in sandbox/prod",
            "Add pre-check decision on org type/config before updates",
            "Release Manager",
            "Open",
        ],
        [
            "R-005",
            "Observability",
            "No explicit fault path or retry guidance",
            "Medium",
            "Medium",
            "Troubleshooting is harder for support teams",
            "Add fault connectors + custom log object entry",
            "Support Lead",
            "Open",
        ],
    ]
    for r in rows:
        ws.append(r)
    style_header(ws)
    add_filter(ws)
    auto_width(ws)
    fills = {
        "High": PatternFill("solid", fgColor="EA9999"),
        "Medium": PatternFill("solid", fgColor="FFE599"),
        "Low": PatternFill("solid", fgColor="B6D7A8"),
    }
    for r in range(2, ws.max_row + 1):
        level = ws[f"D{r}"].value
        ws[f"D{r}"].fill = fills.get(level, PatternFill("solid", fgColor="D9D9D9"))

    ws2 = wb.create_sheet("Recommendations")
    ws2.append(["Priority", "Recommendation", "Business Value", "Technical Note"])
    recs = [
        [
            "P1",
            "Replace hardcoded record name matching with custom metadata-driven keys",
            "Improves reliability across org variants",
            "Move filter constants to `CustomMetadata` and reference via Get Records",
        ],
        [
            "P1",
            "Introduce fault connectors and user-facing error screen",
            "Reduces support ticket triage time",
            "Capture `$Flow.FaultMessage` into a log object",
        ],
        [
            "P2",
            "Add one-time execution guard for org setup",
            "Prevents accidental reruns",
            "Persist setup completion flag in custom setting/metadata",
        ],
    ]
    for r in recs:
        ws2.append(r)
    style_header(ws2)
    add_filter(ws2)
    auto_width(ws2)

    wb.save(BASE / "Risks_and_Recommendations.xlsx")


make_inventory()
make_object_field_impact()
make_dependency_map()
make_integration_map()
make_risks()

